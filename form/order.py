from os import getcwd, path as pathh
from collections import namedtuple
from decimal import Decimal
from PyQt5.QtWidgets import QMainWindow, QMessageBox, QTableWidgetItem, QDialog, QFileDialog
from PyQt5.uic import loadUi
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import Qt, QDate
from pony.orm import *
from my_class.orm_class import Order, OrderPosition, PaymentMethod, ShippingMethod, Client, SupplyPosition, PreOrderPosition, Parts
from form import parts, clients
from form.templates import table
from function.str_to import str_to_decimal
from function.moneyfmt import moneyfmt
import openpyxl
from copy import copy
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill
from openpyxl.drawing.image import Image
from openpyxl.worksheet.pagebreak import Break
import num2t4ru


COLOR_WINDOW = "188, 143, 143"


class OrderList(table.TableList):
    def set_settings(self):
        self.setWindowTitle("Заказы")  # Имя окна
        self.resize(600, 270)
        self.toolBar.setStyleSheet("background-color: rgb(%s);" % COLOR_WINDOW)  # Цвет бара

        self.pb_copy.deleteLater()
        self.pb_other.deleteLater()
        self.pb_filter.deleteLater()

        # Названия колонк (Имя, Длинна)
        self.table_header_name = (("№", 30), ("Клиент", 140), ("Дата заказа", 80), ("Оплата", 140), ("Позиций", 55), ("Сумма", 75))

        self.item = Order  # Класс который будем выводить! Без скобок!

        # сам запрос
        self.query = select((o.id, o.id, o.client.name, o.date, o.payment_method.name, o.value_position, o.sum_all) for o in Order).order_by(-1)

    def ui_add_table_item(self):  # Добавить предмет
        self.add_supply = OrderBrows(self)
        self.add_supply.setWindowModality(Qt.ApplicationModal)
        self.add_supply.show()

    def ui_change_table_item(self, item_id=False):  # изменить элемент
        if item_id:
            item_id = item_id
        else:
            try:
                item_id = self.table_widget.selectedItems()[0].data(5)
            except:
                QMessageBox.information(self, "Ошибка ", "Выделите элемент который хотите изменить", QMessageBox.Ok)
                return False

        self.supply = OrderBrows(self, item_id)
        self.supply.setWindowModality(Qt.ApplicationModal)
        self.supply.show()

    def ui_double_click_table_item(self, item):  # Двойной клик по элементу
        if not self.dc_select:
            self.ui_change_table_item(item.data(5))
        else:
            # что хотим получить ставим всместо 0
            self.main.of_select_order(item.data(5))
            self.close()
            self.destroy()


class OrderBrows(QMainWindow):
    def __init__(self, main=None, order_id=None):
        super(QMainWindow, self).__init__()
        loadUi(getcwd() + '/ui/order.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.toolBar.setStyleSheet("background-color: rgb(%s);" % COLOR_WINDOW)

        self.main = main
        self.id = order_id

        self.PositionOrder = namedtuple('PositionOrder', """sql_id value price sum profit_sum supply_position product order""")

        self.del_position_id = set()  # тут хранятся Id для удаления позиций
        self.del_pre_position_id = set()  # тут хранятся Id для удаления позиций предзаказа

        self.start()

    @db_session
    def start(self):  # Стартовык настройки
        self.tw_position.horizontalHeader().resizeSection(0, 150)
        self.tw_position.horizontalHeader().resizeSection(1, 70)
        self.tw_position.horizontalHeader().resizeSection(2, 50)
        self.tw_position.horizontalHeader().resizeSection(3, 50)
        self.tw_position.horizontalHeader().resizeSection(4, 60)
        self.tw_position.horizontalHeader().resizeSection(5, 60)

        self.de_date.setDate(QDate.currentDate())
        self.de_date_shipping.setDate(QDate.currentDate())
        self.de_date_shipping.setEnabled(False)


        for s in select(s for s in ShippingMethod):
            self.cb_shipping.addItem(s.name, s.id)

        for p in select(p for p in PaymentMethod):
            self.cb_payment.addItem(p.name, p.id)

        if self.id:
            order = Order[self.id]

            self.le_id.setText(str(order.id))
            self.de_date.setDate(order.date)
            self.of_select_client(order.client.id)
            self.cb_shipping.setCurrentText(order.shipping_method.name)
            self.cb_payment.setCurrentText(order.payment_method.name)
            self.le_note.setPlainText(order.note)
            self.le_shipping.setText(str(order.sum_shipping))
            self.le_position_sum.setText(str(order.sum_position))
            self.le_discount_sum.setText(str(order.sum_discount))
            self.le_all_sum.setText(str(order.sum_all))
            self.le_value_position.setText(str(order.value_position))
            self.le_discount_percent.setText(str(order.discount_percent))

            if order.issued:
                self.de_date_shipping.setEnabled(True)
                self.de_date_shipping.setDate(order.date_shipping)
                self.cb_issued.setChecked(True)

            for position in order.order_positions:
                self.tw_position.insertRow(self.tw_position.rowCount())
                for i, position_atr in enumerate((position.supply_position.parts.name, position.supply_position.supply.id,
                                                  position.value, position.price, position.sum)):
                    item = QTableWidgetItem(str(position_atr))
                    if i == 0:  # Вставляем ID записи первую колонку!
                        item.setData(5, position.id)
                    self.tw_position.setItem(self.tw_position.rowCount()-1, i, item)

            for position in order.pre_order_positions:
                self.tw_pre_order.insertRow(self.tw_pre_order.rowCount())
                for i, position_atr in enumerate((position.product.name, position.value, position.price, position.sum)):
                    item = QTableWidgetItem(str(position_atr))
                    if i == 0:  # Вставляем ID записи первую колонку!
                        item.setData(5, position.id)
                    self.tw_pre_order.setItem(self.tw_pre_order.rowCount()-1, i, item)

    def ui_view_doc(self):
        self.doc_list = OrderDocument(self)
        self.doc_list.setModal(True)
        self.doc_list.show()

    def ui_view_client(self):
        self.client = clients.ClientList(self, dc_select=True)
        self.client.setWindowModality(Qt.ApplicationModal)
        self.client.show()

    def ui_add_position(self):  # Добавляем позию товара
        if self.tabWidget.currentIndex() == 0:
            self.position = OrderPositionBrows()
            self.position.setModal(True)
            self.position.show()
            if self.position.exec_() < 1:
                return False

            position = self.position.acc_item
            col = 0
            self.tw_position.insertRow(self.tw_position.rowCount())

            for i in (position.product, position.order, position.value, position.price, position.sum, position.profit_sum):
                item = QTableWidgetItem(str(i))
                if col == 0:  # Вставляем список со значениями только в первую колонку!
                    item.setData(5, position)
                self.tw_position.setItem(self.tw_position.rowCount()-1, col, item)
                col += 1

            self.pb_acc.setEnabled(False)
            self.pb_calc.setStyleSheet("background-color: rgb(255, 61, 44);")

        else:
            self.position = PreOrderPositionBrows()
            self.position.setModal(True)
            self.position.show()
            if self.position.exec_() < 1:
                return False

            position = self.position.acc_item
            col = 0
            self.tw_pre_order.insertRow(self.tw_pre_order.rowCount())

            for i in (position.product, position.value, position.price, position.sum):
                item = QTableWidgetItem(str(i))
                if col == 0:  # Вставляем список со значениями только в первую колонку!
                    item.setData(5, position)
                self.tw_pre_order.setItem(self.tw_pre_order.rowCount()-1, col, item)
                col += 1

    def ui_change_position(self):
        if self.tabWidget.currentIndex() == 0:
            try:
                row = self.tw_position.currentRow()
            except:
                QMessageBox.information(self, "Ошибка", "Выделите позицию которую хотите изменить", QMessageBox.Ok)
                return False
            if row == -1:
                QMessageBox.information(self, "Ошибка", "Выделите позицию которую хотите изменить", QMessageBox.Ok)
                return False

            position = self.tw_position.item(row, 0).data(5)

            self.position = OrderPositionBrows(position)
            self.position.setModal(True)
            self.position.show()
            if self.position.exec_() < 1:
                return False

            position = self.position.acc_item
            col = 0

            for i in (position.product, position.order, position.value, position.price, position.sum, position.profit_sum):
                item = QTableWidgetItem(str(i))
                if col == 0:  # Вставляем список со значениями только в первую колонку!
                    item.setData(5, position)
                self.tw_position.setItem(row, col, item)
                col += 1

            self.pb_acc.setEnabled(False)
            self.pb_calc.setStyleSheet("background-color: rgb(255, 61, 44);")
        else:
            try:
                row = self.tw_pre_order.currentRow()
            except:
                QMessageBox.information(self, "Ошибка", "Выделите позицию которую хотите изменить", QMessageBox.Ok)
                return False
            if row == -1:
                QMessageBox.information(self, "Ошибка", "Выделите позицию которую хотите изменить", QMessageBox.Ok)
                return False

            position = self.tw_pre_order.item(row, 0).data(5)

            self.position = PreOrderPositionBrows(position)
            self.position.setModal(True)
            self.position.show()
            if self.position.exec_() < 1:
                return False

            position = self.position.acc_item
            col = 0

            for i in (position.product, position.value, position.price, position.sum):
                item = QTableWidgetItem(str(i))
                if col == 0:  # Вставляем список со значениями только в первую колонку!
                    item.setData(5, position)
                self.tw_pre_order.setItem(row, col, item)
                col += 1

    def ui_del_position(self):  # Удаляем позицию товара
        if self.tabWidget.currentIndex() == 0:
            result = QMessageBox.question(self, "Удаление", "Точно удалить товар?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if result == 16384:
                try:
                    row = self.tw_position.currentRow()
                except:
                    QMessageBox.critical(self, "Ошибка Удаления", "Выделите товар который хотите удалить", QMessageBox.Ok)
                    return False

                position = self.tw_position.item(row, 0).data(5)  # Смотрим какой тип лежит в дате (кортеж или просто ID)
                if isinstance(position, int):
                    sql_id = position
                elif isinstance(position, str):
                    sql_id = int(position)
                elif isinstance(position, tuple):
                    sql_id = position.sql_id
                else:
                    QMessageBox.critical(self, "Ошибка типа", "Пришел непонятный тип %s" % str(type(position)), QMessageBox.Ok)
                    return False

                if sql_id is None:  # Если нет ID то можно спокойно удалить строку
                    self.tw_position.removeRow(row)
                else:
                    self.del_position_id.add(sql_id)
                    self.tw_position.removeRow(row)

            self.pb_acc.setEnabled(False)
            self.pb_calc.setStyleSheet("background-color: rgb(255, 61, 44);")
        else:
            result = QMessageBox.question(self, "Удаление", "Точно удалить предзаказаный товар?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if result == 16384:
                try:
                    row = self.tw_pre_order.currentRow()
                except:
                    QMessageBox.critical(self, "Ошибка Удаления", "Выделите товар который хотите удалить", QMessageBox.Ok)
                    return False

                position = self.tw_pre_order.item(row, 0).data(5)  # Смотрим какой тип лежит в дате (кортеж или просто ID)
                if isinstance(position, int):
                    sql_id = position
                elif isinstance(position, str):
                    sql_id = int(position)
                elif isinstance(position, tuple):
                    sql_id = position.sql_id
                else:
                    QMessageBox.critical(self, "Ошибка типа", "Пришел непонятный тип %s" % str(type(position)), QMessageBox.Ok)
                    return False

                if sql_id is None:  # Если нет ID то можно спокойно удалить строку
                    self.tw_pre_order.removeRow(row)
                else:
                    self.del_pre_position_id.add(sql_id)
                    self.tw_pre_order.removeRow(row)

    @db_session
    def ui_acc(self):
        value = {
                "date": self.de_date.date().toPyDate(),
                "date_shipping": self.de_date_shipping.date().toPyDate(),
                "sum_shipping": str_to_decimal(self.le_shipping.text()),
                "sum_position": str_to_decimal(self.le_position_sum.text()),
                "sum_discount": str_to_decimal(self.le_discount_sum.text()),
                "sum_all": str_to_decimal(self.le_all_sum.text()),
                "value_position": int(self.le_value_position.text()),
                "discount_percent": str_to_decimal(self.le_discount_percent.text()),
                "note": self.le_note.toPlainText(),
                "client": int(self.le_client.whatsThis()),
                "shipping_method": self.cb_shipping.currentData(),
                "payment_method": self.cb_payment.currentData(),
                "issued": self.cb_issued.isChecked()
                }

        if self.id:
            order = Order[self.id]
            order.set(**value)
        else:
            order = Order(**value)

        for row in range(self.tw_position.rowCount()):  # Добавим или обновим позиции
            position = self.tw_position.item(row, 0).data(5)
            if not isinstance(position, tuple):  # Если не список то ненадо сохранять ничего
                continue

            if position.sql_id is None:  # получим старое проданое кол-во
                last_value = 0
            else:
                last_value = OrderPosition[int(position.sql_id)].value

            value = {
                    "value": position.value,
                    "price": position.price,
                    "sum": position.sum,
                    "supply_position": position.supply_position,
                    "order": order,
                    }

            if position.sql_id is None:
                order.order_positions.add(OrderPosition(**value))
            else:
                OrderPosition[position.sql_id].set(**value)

            # Изменим ко-во на складе
            warehouse_position = SupplyPosition[position.supply_position]
            warehouse_position.warehouse_value += (last_value - position.value)

        for row in range(self.tw_pre_order.rowCount()):  # Добавим или обновим позиции предзаказа
            position = self.tw_pre_order.item(row, 0).data(5)
            if not isinstance(position, tuple):  # Если не список то ненадо сохранять ничего
                continue

            value = {
                    "product": position.product_id,
                    "value": position.value,
                    "price": position.price,
                    "sum": position.sum,
                    "order": order}

            if position.sql_id is None:
                order.pre_order_positions.add(PreOrderPosition(**value))
            else:
                PreOrderPosition[position.sql_id].set(**value)

        for _id in self.del_position_id:
            # Изменим ко-во на складе
            position = OrderPosition[_id]
            position.supply_position.warehouse_value += position.value

            OrderPosition[_id].delete()

        for _id in self.del_pre_position_id:
            PreOrderPosition[_id].delete()

        self.main.ui_update()
        self.close()
        self.destroy()

    def ui_can(self):
        self.close()
        self.destroy()

    @db_session
    def ui_calc(self):  # Подсчет заказа
        shipping_sum = str_to_decimal(self.le_shipping.text())
        discount_percent = str_to_decimal(self.le_discount_percent.text())

        position_sum = 0
        value_position = 0
        profit_sum = 0

        for row in range(self.tw_position.rowCount()):
            position = self.tw_position.item(row, 0).data(5)
            if not isinstance(position, tuple):
                sql_position = OrderPosition[int(position)]  # Получаем позицию из бд
                position = self.PositionOrder(sql_position.id, sql_position.value, sql_position.price, sql_position.sum,
                                              sql_position.supply_position.price_cost, sql_position.supply_position.id,
                                              sql_position.supply_position.parts.name, sql_position.supply_position.supply.id)

            position_sum += position.sum
            value_position += 1
            profit_sum += position.profit_sum

        if position_sum:
            discount_sum = (position_sum / 100) * discount_percent
        else:
            discount_sum = 0

        all_sum = (position_sum - discount_sum) + shipping_sum
        profit_sum = profit_sum - discount_sum

        self.le_position_sum.setText(str(position_sum))
        self.le_discount_sum.setText(str(discount_sum))
        self.le_all_sum.setText(str(all_sum))
        self.le_value_position.setText(str(value_position))
        self.le_profit_sum.setText(str(profit_sum))

        self.pb_acc.setEnabled(True)
        self.pb_calc.setStyleSheet("background-color: rgb(85, 255, 0);")

    @db_session
    def of_select_client(self, client_id):
        client = Client[client_id]
        self.le_client.setText(client.name)
        self.le_client.setWhatsThis(str(client.id))

    @db_session
    def of_ex_score(self, pre_order=False):  # Составляем счет
        path = QFileDialog.getSaveFileName(self, "Сохранение", pathh.expanduser("~/Desktop/"), filter="Excel(*.xlsx)")
        if not path[0]:
            return False

        book = openpyxl.load_workbook(filename='%s/score.xlsx' % (getcwd() + "/templates/order"))
        sheet = book['s1']

        wite = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')
        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        sheet["A9"] = "Счет № %s от %s г." % (self.le_id.text(), self.de_date.date().toString("dd.MM.yyyy"))

        client = Order[self.id].client
        sheet["B13"] = "Плательщик: %s ИНН:%s КПП%s , Адрес: %s" % (client.full_name, client.inn, client.kpp, client.addres_legal)

        all_sum = 0
        row_ex = 16
        value_row = 1
        if not pre_order:  # Если это будет предзаказ то втавлять будем из другой таблицы
            for row in range(self.tw_position.rowCount()):
                sheet["A%s" % row_ex] = str(value_row)
                sheet["B%s" % row_ex] = self.tw_position.item(row, 0).text()
                sheet["C%s" % row_ex] = float(self.tw_position.item(row, 2).text())
                sheet["D%s" % row_ex] = "шт."
                sheet["E%s" % row_ex] = float(self.tw_position.item(row, 3).text())
                sheet["F%s" % row_ex] = float(self.tw_position.item(row, 4).text())

                all_sum += Decimal(self.tw_position.item(row, 4).text())
                row_ex += 1
                value_row += 1
        else:
            for row in range(self.tw_pre_order.rowCount()):
                sheet["A%s" % row_ex] = str(value_row)
                sheet["B%s" % row_ex] = self.tw_pre_order.item(row, 0).text()
                sheet["C%s" % row_ex] = float(self.tw_pre_order.item(row, 1).text())
                sheet["D%s" % row_ex] = "шт."
                sheet["E%s" % row_ex] = float(self.tw_pre_order.item(row, 2).text())
                sheet["F%s" % row_ex] = float(self.tw_pre_order.item(row, 3).text())

                all_sum += Decimal(self.tw_pre_order.item(row, 3).text())
                row_ex += 1
                value_row += 1

        for row in sheet.iter_rows(min_row=16, max_col=6, max_row=row_ex-1):
            for cell in row:
                cell.border = border_all

        # низ счета
        sheet2 = book['s2']
        sheet2['F1'] = all_sum
        sheet2['F3'] = all_sum
        sheet2['A5'] = "Всего наименований %s, на сумму %s" % (value_row, all_sum)

        for row in sheet2.iter_rows(min_row=1, max_col=6, max_row=14):
            for cell in row:
                sheet["%s%s" % (cell.column, row_ex)] = cell.value
                sheet.row_dimensions[row_ex].height = sheet2.row_dimensions[cell.row].height
                sheet["%s%s" % (cell.column, row_ex)].fill = wite
                if cell.has_style:
                    sheet["%s%s" % (cell.column, row_ex)].border = copy(cell.border)
                    sheet["%s%s" % (cell.column, row_ex)].font = copy(cell.font)
            row_ex += 1
        book.remove(sheet2)


        # Вставляем черные квадраты
        self.path_templates = getcwd() + '/templates/order'

        sheet = book["s1"]
        img = Image('%s/logo.jpg' % self.path_templates)
        sheet.add_image(img, 'A1')

        book.save(path[0])

    @db_session
    def of_ex_torg12(self, pre_order=False):
        path = QFileDialog.getSaveFileName(self, "Сохранение", pathh.expanduser("~/Desktop/"), filter="Excel(*.xlsx)")
        if not path[0]:
            return False

        border_all = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        border_all_big = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))

        font_8 = Font(name="Arial", size=8)

        ald_center = Alignment(horizontal="center")
        ald_right = Alignment(horizontal="right")

        wite = PatternFill(start_color='ffffff', end_color='ffffff', fill_type='solid')

        book = openpyxl.load_workbook(filename='%s/torg12.xlsx' % (getcwd() + "/templates/order"))
        sheet = book['Отчет']

        sheet.oddHeader.right.text = "Продолжение накладной № %s от %s г." % (self.le_id.text(), self.de_date.date().toString("dd.MM.yyyy"))
        sheet.oddHeader.right.size = 7

        client = Order[self.id].client

        client_text = client.full_name + " ИНН " + str(client.inn)
        if client.kpp:
            client_text += " КПП " + str(client.kpp)
        client_text += ", " + client.addres_legal + ",р/с " + str(client.account) +\
                       " в " + client.bank + " к/с " + str(client.corres_account) + " БИК " + str(client.bik)
        sheet["C8"] = client_text
        sheet["C12"] = client_text

        sheet["C14"] = "Счета № %s от %s" % (self.le_id.text(), self.de_date.date().toString("dd.MM.yyyy"))

        # заполнение середины
        sheet["G17"] = self.le_id.text()
        sheet["T13"] = self.le_id.text()
        sheet["I17"] = self.de_date.date().toString("dd.MM.yyyy")
        sheet["T14"] = self.de_date.date().toString("dd.MM.yyyy")
        sheet["G17"].border = border_all_big
        sheet["I17"].border = border_all_big

        all_value = 0
        all_sum = 0
        all_position = 0

        list_all = 1
        row_break = 14
        row_ex = 22

        num = 1
        if not pre_order:  # Если это будет предзаказ то втавлять будем из другой таблицы
            for row in range(self.tw_position.rowCount()):
                sheet.merge_cells("B%s:D%s" % (row_ex, row_ex))
                sheet.merge_cells("L%s:M%s" % (row_ex, row_ex))
                sheet.merge_cells("N%s:O%s" % (row_ex, row_ex))
                sheet.merge_cells("P%s:Q%s" % (row_ex, row_ex))
                sheet.merge_cells("R%s:S%s" % (row_ex, row_ex))
                sheet.merge_cells("T%s:U%s" % (row_ex, row_ex))

                sheet["A%s" % row_ex] = num
                sheet["B%s" % row_ex] = self.tw_position.item(row, 0).text()
                sheet["B%s" % row_ex].alignment = Alignment(wrapText=True)
                # sheet["B%s" % row_ex].font = font_8
                sheet["E%s" % row_ex] = ""
                sheet["F%s" % row_ex] = "шт."
                sheet["G%s" % row_ex] = "796"
                sheet["H%s" % row_ex] = "кор."
                sheet["I%s" % row_ex] = ""
                sheet["J%s" % row_ex] = ""
                sheet["L%s" % row_ex] = float(self.tw_position.item(row, 2).text())
                sheet["N%s" % row_ex] = moneyfmt(self.tw_position.item(row, 3).text())
                sheet["N%s" % row_ex].alignment = ald_right
                sheet["P%s" % row_ex] = ""
                sheet["R%s" % row_ex] = "-"
                sheet["T%s" % row_ex] = "-"
                sheet["V%s" % row_ex] = moneyfmt(self.tw_position.item(row, 4).text())
                sheet["V%s" % row_ex].alignment = ald_right

                all_value += float(self.tw_position.item(row, 2).text())
                all_sum += Decimal(self.tw_position.item(row, 4).text())

                sheet.row_dimensions[row_ex].height = 23

                if row_break == 25 and self.tw_position.rowCount() > 16:
                    sheet.page_breaks.append(Break(row_ex))
                    list_all += 1
                    row_break = 0

                row_break += 1
                row_ex += 1
                num += 1
        else:
            for row in range(self.tw_pre_order.rowCount()):
                sheet.merge_cells("B%s:D%s" % (row_ex, row_ex))
                sheet.merge_cells("L%s:M%s" % (row_ex, row_ex))
                sheet.merge_cells("N%s:O%s" % (row_ex, row_ex))
                sheet.merge_cells("P%s:Q%s" % (row_ex, row_ex))
                sheet.merge_cells("R%s:S%s" % (row_ex, row_ex))
                sheet.merge_cells("T%s:U%s" % (row_ex, row_ex))

                sheet["A%s" % row_ex] = num
                sheet["B%s" % row_ex] = self.tw_pre_order.item(row, 0).text()
                sheet["B%s" % row_ex].alignment = Alignment(wrapText=True)
                # sheet["B%s" % row_ex].font = font_8
                sheet["E%s" % row_ex] = ""
                sheet["F%s" % row_ex] = "шт."
                sheet["G%s" % row_ex] = "796"
                sheet["H%s" % row_ex] = "кор."
                sheet["I%s" % row_ex] = ""
                sheet["J%s" % row_ex] = ""
                sheet["L%s" % row_ex] = float(self.tw_pre_order.item(row, 1).text())
                sheet["N%s" % row_ex] = moneyfmt(self.tw_pre_order.item(row, 2).text())
                sheet["N%s" % row_ex].alignment = ald_right
                sheet["P%s" % row_ex] = ""
                sheet["R%s" % row_ex] = "-"
                sheet["T%s" % row_ex] = "-"
                sheet["V%s" % row_ex] = moneyfmt(self.tw_pre_order.item(row, 3).text())
                sheet["V%s" % row_ex].alignment = ald_right

                all_value += float(self.tw_pre_order.item(row, 1).text())
                all_sum += Decimal(self.tw_pre_order.item(row, 3).text())

                sheet.row_dimensions[row_ex].height = 23

                if row_break == 25 and self.tw_pre_order.rowCount() > 16:
                    sheet.page_breaks.append(Break(row_ex))
                    list_all += 1
                    row_break = 0

                row_break += 1
                row_ex += 1
                num += 1


        if row_break + 7 > 25:
            sheet.page_breaks.append(Break(row_ex-4))
            list_all += 1

        # Заполняем сумму
        sheet["I%s" % row_ex] = "Всего по накладной"
        sheet["I%s" % row_ex].alignment = ald_right

        sheet["J%s" % row_ex] = "X"
        sheet["J%s" % row_ex].alignment = ald_center

        sheet["K%s" % row_ex] = "X"
        sheet["K%s" % row_ex].alignment = ald_center

        sheet.merge_cells("L%s:M%s" % (row_ex, row_ex))
        sheet["L%s" % row_ex] = all_value
        sheet["L%s" % row_ex].alignment = ald_right

        sheet.merge_cells("N%s:O%s" % (row_ex, row_ex))
        sheet["N%s" % row_ex] = "X"
        sheet["N%s" % row_ex].alignment = ald_center

        sheet.merge_cells("P%s:Q%s" % (row_ex, row_ex))
        sheet["P%s" % row_ex] = "X"
        sheet["P%s" % row_ex].alignment = ald_center

        sheet.merge_cells("R%s:S%s" % (row_ex, row_ex))
        sheet["R%s" % row_ex] = "X"
        sheet["R%s" % row_ex].alignment = ald_center

        sheet.merge_cells("T%s:U%s" % (row_ex, row_ex))
        sheet["T%s" % row_ex] = "X"
        sheet["T%s" % row_ex].alignment = ald_center

        sheet["V%s" % row_ex] = moneyfmt(all_sum)
        sheet["V%s" % row_ex].alignment = ald_right

        for row in sheet.iter_rows(min_row=row_ex, min_col=10, max_col=22):
            for cell in row:
                cell.border = border_all

        row_ex += 1

        # Формируем шапку
        for row in sheet.iter_rows(min_row=19, max_col=22, max_row=21):
            for cell in row:
                    cell.border = border_all_big

        for row in sheet.iter_rows(min_row=4, min_col=20, max_row=5):
            for cell in row:
                cell.border = border_all_big

        for row in sheet.iter_rows(min_row=17, min_col=7, max_col=10, max_row=17):
            for cell in row:
                cell.border = border_all_big

        for row in sheet.iter_rows(min_row=13, min_col=17, max_col=19, max_row=16):
            for cell in row:
                cell.border = border_all

        # Формируем шапку-правую колонку
        for row in sheet.iter_rows(min_row=13, min_col=20, max_col=22, max_row=14):
            for cell in row:
                cell.border = border_all
        sheet["T13"].border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet["T14"].border = Border(left=Side(style='medium'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet["V13"].border = Border(left=Side(style='thin'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))
        sheet["V14"].border = Border(left=Side(style='thin'), right=Side(style='medium'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Формируем границы таблицы
        for row in sheet.iter_rows(min_row=22, max_col=22, max_row=row_ex-2):
            for cell in row:
                cell.border = border_all

        sheet2 = book['Низ']

        for row in sheet2.iter_rows(min_row=1, max_col=22, max_row=17):
            for cell in row:
                sheet["%s%s" % (cell.column, row_ex)] = cell.value
                sheet.row_dimensions[row_ex].height = sheet2.row_dimensions[cell.row].height
                if cell.has_style:
                    sheet["%s%s" % (cell.column, row_ex)].border = copy(cell.border)
                    sheet["%s%s" % (cell.column, row_ex)].font = copy(cell.font)
                    sheet["%s%s" % (cell.column, row_ex)].fill = wite

            row_ex += 1

        sheet["I%s" % (row_ex - 17)] = list_all

        # Числа прописью
        int_units = ((u'рубль', u'рубля', u'рублей'), 'm')
        exp_units = ((u'копейка', u'копейки', u'копеек'), 'f')
        sheet["D%s" % (row_ex-16)] = num2t4ru.num2text(num-1)
        # sheet["D%s" % (row_ex-13)] = num2t4ru.num2text(all_position)
        sheet["D%s" % (row_ex-9)] = num2t4ru.decimal2text(Decimal(str(all_sum)), int_units=int_units, exp_units=exp_units)

        book.remove(sheet2)

        book.save(path[0])


class OrderPositionBrows(QDialog):
    def __init__(self, position=None):
        super(OrderPositionBrows, self).__init__()
        loadUi(getcwd() + '/ui/order_position.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.widget.setStyleSheet("background-color: rgb(%s);" % COLOR_WINDOW)

        self.position = position

        self.sql_id = None  # ID sql позиции
        self.acc_item = None  # переменная для хранения итогово списка
        self.warehouse_value = None  # переменная для записи кол-ва товара на складе

        self.PositionOrder = namedtuple('PositionOrder', """sql_id value price sum profit_sum supply_position product order""")   # Переменная хранения позиции

        self.start()

    @db_session
    def start(self):
        if self.position:
            if isinstance(self.position, int):
                position = OrderPosition[self.position]

                self.position = self.PositionOrder(position.id, position.value, position.price, position.sum, " ",
                                                   position.supply_position.id, position.supply_position.parts.name,
                                                   position.supply_position.supply.id)

                self.pb_catalog.setEnabled(False)  # Если это сохранеая позиция то нельзя изменить товар.

            self.sql_id = self.position.sql_id

            supply = SupplyPosition[self.position.supply_position]
            self.warehouse_value = supply.warehouse_value
            self.le_cost_price.setText(str(supply.price_cost))

            if self.sql_id:  # Если позиция уже сохранена то к ко-ву на складе прибавить кол-во проданого
                self.warehouse_value += self.position.value

            self.le_product.setText(str(self.position.product))
            self.le_warehouse.setWhatsThis(str(self.position.supply_position))
            self.le_warehouse.setText(str(self.position.order))
            self.le_value.setText(str(self.position.value))
            self.le_price.setText(str(self.position.price))
            self.le_sum.setText(str(self.position.sum))

            self.ui_calc_sum()
            self.calc_profit()

        else:
            self.warehouse_value = 0

    def ui_view_catalog(self):
        self.parts = parts.PartsCatalog(self, select_warehouse=True)
        self.parts.setWindowModality(Qt.ApplicationModal)
        self.parts.show()

    def ui_acc(self):
        if not self.le_warehouse.text():
            QMessageBox.information(self, "Ошибка", "Не выбран товар", QMessageBox.Ok)
            return False

        self.acc_item = self.PositionOrder(self.sql_id,
                                           str_to_decimal(self.le_value.text()),
                                           str_to_decimal(self.le_price.text()),
                                           str_to_decimal(self.le_sum.text()),
                                           str_to_decimal(self.le_profit_sum.text()),
                                           int(self.le_warehouse.whatsThis()),
                                           self.le_product.text(),
                                           self.le_warehouse.text(),
                                           )

        if self.position == self.acc_item:  # Если позиция не изменилась то дклаем закрытие окна
            print("нет изменений в пизиции")
            self.ui_can()
            return False

        # Проверим список на заполненость! Кроме первого значения. Там может быть None если это новая позиция
        for item in tuple(self.acc_item)[1:]:
            if not item:
                QMessageBox.information(self, "Ошибка заполнения", "Что то не заполнено", QMessageBox.Ok)
                return False

        self.done(1)
        self.close()
        self.destroy()

    def ui_can(self):
        self.done(0)
        self.close()
        self.destroy()

    def ui_calc_value(self):  # Смотрит сколько товара есть на селаде и сравнивает с кол-вом
        value = str_to_decimal(self.le_value.text())

        if value and value > self.warehouse_value:
            self.le_value.setText("")

    def ui_calc_price(self):  # посчитать цену от суммы
        value = str_to_decimal(self.le_value.text())
        sum = str_to_decimal(self.le_sum.text())
        cost_sum = str_to_decimal(self.le_cost_sum.text())

        if value and sum:
            self.le_price.setText(str(round(sum / value, 2)))

        if value and cost_sum:
            self.le_cost_price.setText(str(round(sum / cost_sum, 2)))

        self.calc_profit()

    def ui_calc_sum(self):  # Считаем сумму цены продавца и рублей и с наценкой
        value = str_to_decimal(self.le_value.text())
        price = str_to_decimal(self.le_price.text())
        cost_price = str_to_decimal(self.le_cost_price.text())

        if value and price:
            self.le_sum.setText(str(round(value * price, 2)))

        if value and cost_price:
            self.le_cost_sum.setText(str(round(value * cost_price, 2)))

        self.calc_profit()

    def calc_profit(self):  # Высчитываем прибыль
        price = str_to_decimal(self.le_price.text())
        cost_price = str_to_decimal(self.le_cost_price.text())

        sum = str_to_decimal(self.le_sum.text())
        cost_sum = str_to_decimal(self.le_cost_sum.text())

        if price and cost_price:
            self.le_price_profit.setText(str(round(price - cost_price, 2)))

        if sum and cost_sum:
            self.le_profit_sum.setText(str(round(sum - cost_sum, 2)))

    @db_session
    def of_tree_select_catalog_warehouse(self, _id):
        supply_position = SupplyPosition[_id]

        self.le_product.setText(supply_position.parts.name)
        self.le_warehouse.setText(str(supply_position.supply.id))
        self.le_warehouse.setWhatsThis(str(supply_position.id))

        self.le_price.setText(str(supply_position.price_sell))
        self.le_cost_price.setText(str(supply_position.price_cost))

        self.warehouse_value = supply_position.warehouse_value

        self.le_value.setFocus()

        self.ui_calc_sum()
        self.calc_profit()


class OrderDocument(QDialog):
    def __init__(self, main):
        super(OrderDocument, self).__init__()
        loadUi(getcwd() + '/ui/order_doc.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/order_doc.ico"))

        self.main = main

    def ui_acc(self):
        if self.lw_main.selectedItems()[0].text() == "Накладная":
            pre_order = self.cb_pre_order.isChecked()
            self.main.of_ex_torg12(pre_order=pre_order)

        elif self.lw_main.selectedItems()[0].text() == "Счет":
            pre_order = self.cb_pre_order.isChecked()
            self.main.of_ex_score(pre_order=pre_order)

        self.close()
        self.destroy()

    def ui_can(self):
        self.close()
        self.destroy()


class PreOrderPositionBrows(QDialog):
    def __init__(self, position=None):
        super(PreOrderPositionBrows, self).__init__()
        loadUi(getcwd() + '/ui/pre_order_position.ui', self)
        self.setWindowIcon(QIcon(getcwd() + "/images/icon.ico"))
        self.widget.setStyleSheet("background-color: rgb(%s);" % COLOR_WINDOW)

        self.PositionPreOrder = namedtuple('PositionCost', 'sql_id, product_id product value price sum')

        self.position = position

        self.sql_id = None  # переменная для хранения и вставки sql ID строки
        self.acc_item = None  # переменная для хранения итогово списка

        self.start()

    @db_session
    def start(self):
        # Проверим что пришло, список или Id строки
        if isinstance(self.position, tuple):
            self.sql_id = self.position.sql_id
            self.le_parts.setWhatsThis(str(self.position.product_id))
            self.le_parts.setText(str(self.position.product))
            self.le_value.setText(str(self.position.value))
            self.le_price.setText(str(self.position.price))
            self.le_sum.setText(str(self.position.sum))

        elif isinstance(self.position, int):
            self.sql_id = self.position  # Записываем Id строки
            sql_position = PreOrderPosition[self.sql_id]  # Получаем позицию из бд
            self.position = self.PositionPreOrder(self.sql_id, sql_position.product.id, sql_position.product.name,  # Составляем список для последующего сравнения
                                              sql_position.value, sql_position.price, sql_position.sum)
            self.le_parts.setWhatsThis(str(self.position.product_id))
            self.le_parts.setText(str(self.position.product))
            self.le_value.setText(str(self.position.value))
            self.le_price.setText(str(self.position.price))
            self.le_sum.setText(str(self.position.sum))

        elif self.position is None:  # Новая позиция
            pass

        else:
            QMessageBox.information(self, "Ошибка типа", "Пришел непонятный тип %s" % str(type(self.position)), QMessageBox.Ok)
            return False

    def ui_view_parts(self):
        self.parts = parts.PartsCatalog(self, select_product=True)
        self.parts.setWindowModality(Qt.ApplicationModal)
        self.parts.show()

    def ui_acc(self):
        if not self.le_parts.text():
            QMessageBox.information(self, "Ошибка", "Не выбрана затрата", QMessageBox.Ok)
            return False

        self.acc_item = self.PositionPreOrder(self.sql_id,
                                              int(self.le_parts.whatsThis()),
                                              self.le_parts.text(),
                                              str_to_decimal(self.le_value.text()),
                                              str_to_decimal(self.le_price.text()),
                                              str_to_decimal(self.le_sum.text()))

        if self.acc_item == self.position:  # Если позиция не изменилась то дклаем закрытие окна
            print("нет изменений в прочих расходах")
            self.ui_can()
            return False

        if not self.acc_item.value:
            QMessageBox.information(self, "Ошибка", "Что то не так с кол-вом", QMessageBox.Ok)
            return False

        if not self.acc_item.price:
            QMessageBox.information(self, "Ошибка", "Что то не так с ценой", QMessageBox.Ok)
            return False

        if not self.acc_item.sum:
            QMessageBox.information(self, "Ошибка", "Что то не так с суммой", QMessageBox.Ok)
            return False

        self.done(1)
        self.close()
        self.destroy()

    def ui_can(self):
        self.done(0)
        self.close()
        self.destroy()

    def ui_calc_sum(self):
        value = str_to_decimal(self.le_value.text())
        price = str_to_decimal(self.le_price.text())

        if value and price:
            self.le_sum.setText(str(round(value*price, 2)))
        else:
            self.le_sum.setText("ОШИБКА")

    def ui_calc_pcs(self):
        value = str_to_decimal(self.le_value.text())
        sum = str_to_decimal(self.le_sum.text())

        if value and sum:
            self.le_price.setText(str(round(sum/value, 2)))
        else:
            self.le_price.setText("ОШИБКА")

    @db_session
    def of_tree_select_catalog_product(self, _id):
        position = Parts[_id]

        self.le_parts.setText(position.name)
        self.le_parts.setWhatsThis(str(position.id))



